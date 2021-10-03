from openpyxl import load_workbook, Workbook
from update_sheet.helper.copy_helper import copy_sheet
from update_sheet.update_cairn import update_reserve, update_bilan, update_fin_de_mois
from update_sheet.helper.sheetname_helper import get_dates_for_month
from app_config import template_folder, output_folder

# creates new file for update_sheet usin a template
def create_cairn(year, month):
    doc_type = 'Cairn'
    print(f'Start creating new file for {doc_type} for {year}-{month}')
    # define all sheet additional to the 'day'-sheets
    fin_de_mois = 'FIN DE MOIS'
    bilan = 'BILAN'
    sheets_to_copy = [fin_de_mois, bilan]

    # load template
    wb = load_workbook(filename=f'{template_folder}/Suivi des cairns 26.xlsx')
    # create new workbook
    wb_new = Workbook()

    # read sheet template from template workbook
    # get names of all day in month
    sheetnames_date = get_dates_for_month(year, month)

    # create sheet for every day in month
    for index, sheet_name in enumerate(sheetnames_date):
        wb_new.create_sheet(sheet_name)
        current_sheet = wb_new[sheet_name]
        # update the ouverture table in every sheet but the first one
        if index != 0:
            # ouverture table references to data in sheet of day before
            copy_sheet(wb['B'], current_sheet)
            update_reserve(current_sheet, sheetnames_date[index-1])
        else:
            copy_sheet(wb['A'], current_sheet)

    for sheet_name in sheets_to_copy:
        wb_new.create_sheet(sheet_name)
        copy_sheet(wb[sheet_name], wb_new[sheet_name])

    update_fin_de_mois(wb_new[fin_de_mois], year, month)
    update_bilan(wb_new[bilan], year, month)

    # remove default sheet 'Sheet' from workbook
    wb_new.remove(wb_new['Sheet'])
    # save new file
    filename = f'{output_folder}{year}.{month}-Suivi des cairns.xlsx'
    wb_new.save(filename)
    print(f'Created new file {doc_type}[{filename}]')
