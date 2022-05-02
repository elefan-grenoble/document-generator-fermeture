import datetime

from openpyxl import load_workbook, Workbook

from app_config import template_folder
from update_sheet.helper.copy_helper import copy_sheet
from update_sheet.helper.sheetname_helper import get_dates_for_month
from update_sheet.update_cb import update_bilan


# creates new file for update_sheet usin a template
def create_cb(year, month, output_folder):
    year_month = datetime.date(year, month, 1).strftime("%Y.%m")
    filename = f'{output_folder}{year_month} - Suivi des achats en carte bancaire.xlsx'

    doc_type = 'CB'
    print(f'Start creating new file for {doc_type} for {year}-{month} [{filename}]')
    # define all sheet additional to the 'day'-sheets
    bilan = 'BILAN'
    sheets_to_copy = [bilan]

    # load template
    wb = load_workbook(filename=f'{template_folder}Suivi des achats en carte bancaire.xlsx')
    # create new workbook
    wb_new = Workbook()

    # read sheet template from template workbook
    # get names of all day in month
    sheetnames_date = get_dates_for_month(year, month)

    # create sheet for every day in month
    for index, sheet_name in enumerate(sheetnames_date):
        wb_new.create_sheet(sheet_name)
        current_sheet = wb_new[sheet_name]
        # update the ouverture table in every sheet but the first one
        copy_sheet(wb['A'], current_sheet)

    for sheet_name in sheets_to_copy:
        wb_new.create_sheet(sheet_name)
        copy_sheet(wb[sheet_name], wb_new[sheet_name])

    update_bilan(wb_new[bilan], year, month)

    # remove default sheet 'Sheet' from workbook
    wb_new.remove(wb_new['Sheet'])
    # save new file
    wb_new.save(filename)
    print(f'Created new file {doc_type}[{filename}]')
